#!/usr/bin/env python3
"""Generate SF Porta Potty financial model (.xlsx) with live Excel formulas.

Architecture:
  - Assumptions sheet: raw input values (the ONLY place numbers are typed)
  - 3-Year Projection: formulas reference Assumptions; helper rows keep each formula short
  - Dashboard: individual line items per scenario, each a simple formula
  - Return Analysis: references Projection; uses FV() for loan balance
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO, "output", "sf-porta-potty")
OUT_FILE = os.path.join(OUT_DIR, "030326-financial-model.xlsx")

# ── Styles ───────────────────────────────────────────────────────────────────
NAVY = "1A2744"
GOLD = "C9A84C"
GREEN_HEX = "27AE60"
RED_HEX = "E74C3C"

header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
label_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", bold=True, size=11)
title_font = Font(name="Calibri", bold=True, size=16, color=NAVY)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=NAVY)
note_font = Font(name="Calibri", size=10, italic=True, color="666666")
helper_font = Font(name="Calibri", size=10, italic=True, color="888888")
editable_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
section_fill = PatternFill(start_color="E8EAF0", end_color="E8EAF0", fill_type="solid")
helper_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

USD = '#,##0'
PCT = '0.0%'
RATIO = '0.00"x"'

green_font_cf = Font(name="Calibri", bold=True, color=GREEN_HEX)
red_font_cf = Font(name="Calibri", bold=True, color=RED_HEX)
green_fill_cf = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
red_fill_cf = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

# ── Assumption row map (row numbers on the Assumptions sheet, data starts row 4) ──
A_UNITS_Y1 = 4;  A_UNITS_Y2 = 5;  A_UNITS_Y3 = 6
A_TRUCKS_Y1 = 7; A_TRUCKS_Y2 = 8; A_TRUCKS_Y3 = 9
A_RATE = 10;     A_EVENT_PCT = 11; A_EVENT_PREMIUM = 12
A_OWNER = 13;    A_DRIVER_Y1 = 14; A_DRIVER_Y2 = 15; A_DRIVER_Y3 = 16
A_CHEM = 17;     A_FUEL = 18
A_YARD = 19;     A_INS = 20;  A_OFFICE = 21; A_MKTG = 22
A_PRICE = 23;    A_LOAN = 24; A_RATE_SBA = 25; A_TERM = 26; A_DOWN = 27

YEAR_ROWS = {
    1: {"units": A_UNITS_Y1, "trucks": A_TRUCKS_Y1, "driver": A_DRIVER_Y1},
    2: {"units": A_UNITS_Y2, "trucks": A_TRUCKS_Y2, "driver": A_DRIVER_Y2},
    3: {"units": A_UNITS_Y3, "trucks": A_TRUCKS_Y3, "driver": A_DRIVER_Y3},
}


def a(row, col="D"):
    """Absolute reference to Assumptions sheet cell."""
    return f"Assumptions!${col}${row}"


def p(cell):
    """Reference to 3-Year Projection sheet cell."""
    return f"'3-Year Projection'!{cell}"


# ── Scenario input values ────────────────────────────────────────────────────
SCENARIOS = {
    "Conservative": {
        "units_y1": 120, "units_y2": 140, "units_y3": 165,
        "trucks_y1": 1, "trucks_y2": 1, "trucks_y3": 2,
        "rate": 185, "event_pct": 0.05, "event_premium": 350,
        "owner_comp": 50000,
        "driver_y1": 0, "driver_y2": 45000, "driver_y3": 55000,
        "chem": 20, "fuel": 1300,
        "yard": 3500, "insurance": 28000, "office": 12000, "marketing": 6000,
        "purchase": 250000, "loan": 250000, "sba_rate": 0.09, "term": 10, "down": 0.10,
    },
    "Base": {
        "units_y1": 150, "units_y2": 175, "units_y3": 210,
        "trucks_y1": 1, "trucks_y2": 2, "trucks_y3": 2,
        "rate": 190, "event_pct": 0.05, "event_premium": 400,
        "owner_comp": 65000,
        "driver_y1": 0, "driver_y2": 52000, "driver_y3": 65000,
        "chem": 20, "fuel": 1300,
        "yard": 5000, "insurance": 33000, "office": 15000, "marketing": 9000,
        "purchase": 500000, "loan": 500000, "sba_rate": 0.09, "term": 10, "down": 0.10,
    },
    "Aggressive": {
        "units_y1": 200, "units_y2": 250, "units_y3": 320,
        "trucks_y1": 2, "trucks_y2": 3, "trucks_y3": 4,
        "rate": 200, "event_pct": 0.05, "event_premium": 450,
        "owner_comp": 80000,
        "driver_y1": 50000, "driver_y2": 100000, "driver_y3": 150000,
        "chem": 20, "fuel": 1300,
        "yard": 7500, "insurance": 43000, "office": 18000, "marketing": 15000,
        "purchase": 750000, "loan": 750000, "sba_rate": 0.09, "term": 10, "down": 0.10,
    },
}

ASSUMPTION_KEYS = [
    "units_y1", "units_y2", "units_y3",
    "trucks_y1", "trucks_y2", "trucks_y3",
    "rate", "event_pct", "event_premium",
    "owner_comp", "driver_y1", "driver_y2", "driver_y3",
    "chem", "fuel",
    "yard", "insurance", "office", "marketing",
    "purchase", "loan", "sba_rate", "term", "down",
]

ASSUMPTION_LABELS = [
    ("Fleet", "Units Year 1"), ("", "Units Year 2"), ("", "Units Year 3"),
    ("", "Trucks Year 1"), ("", "Trucks Year 2"), ("", "Trucks Year 3"),
    ("Revenue", "Monthly Rate ($/unit)"), ("", "Event Mix (%)"), ("", "Event Premium ($/unit/mo)"),
    ("Labor", "Owner Compensation ($/yr)"),
    ("", "Driver Cost Year 1 ($/yr)"), ("", "Driver Cost Year 2 ($/yr)"), ("", "Driver Cost Year 3 ($/yr)"),
    ("Operations", "Chemical Cost ($/unit/mo)"), ("", "Fuel + Maint ($/truck/mo)"),
    ("Overhead", "Yard Lease ($/mo)"), ("", "Insurance ($/yr)"),
    ("", "Office & Misc ($/yr)"), ("", "Marketing ($/yr)"),
    ("Financing", "Purchase Price ($)"), ("", "SBA Loan Amount ($)"),
    ("", "SBA Interest Rate"), ("", "SBA Term (years)"), ("", "Down Payment (%)"),
]

ASSUMPTION_FMTS = [
    USD, USD, USD, USD, USD, USD,
    USD, PCT, USD,
    USD, USD, USD, USD,
    USD, USD,
    USD, USD, USD, USD,
    USD, USD, PCT, USD, PCT,
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def set_formula(ws, row, col, formula, fmt=None, bold=False, is_helper=False):
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.font = helper_font if is_helper else (bold_font if bold else label_font)
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt
    if is_helper:
        cell.fill = helper_fill
    return cell


def row_label(ws, row, text, bold=False, section=False, helper=False):
    cell = ws.cell(row=row, column=1, value=text)
    if section:
        cell.font = bold_font
        cell.fill = section_fill
    elif helper:
        cell.font = helper_font
        cell.fill = helper_fill
    elif bold:
        cell.font = bold_font
    else:
        cell.font = label_font
    cell.border = thin_border


def section_row(ws, row, text, max_col):
    row_label(ws, row, text, section=True)
    for c in range(2, max_col + 1):
        ws.cell(row=row, column=c).fill = section_fill
        ws.cell(row=row, column=c).border = thin_border


def auto_width(ws, min_w=12, max_w=35):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        best = min_w
        for cell in col_cells:
            if cell.value and not str(cell.value).startswith("="):
                best = max(best, min(len(str(cell.value)) + 3, max_w))
        ws.column_dimensions[letter].width = best


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: ASSUMPTIONS  (raw values only — unchanged)
# ═══════════════════════════════════════════════════════════════════════════════
def build_assumptions(ws):
    ws.sheet_properties.tabColor = GOLD

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Model Assumptions — Edit These Cells to Update All Sheets"
    ws["A1"].font = title_font

    ws.merge_cells("A2:E2")
    ws["A2"].value = "Yellow cells are inputs. All other sheets auto-recalculate from these values."
    ws["A2"].font = note_font

    row = 3
    for i, h in enumerate(["Category", "Variable", "Conservative", "Base", "Aggressive"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 5)

    for i, (key, (cat_label, var_label), fmt) in enumerate(
            zip(ASSUMPTION_KEYS, ASSUMPTION_LABELS, ASSUMPTION_FMTS)):
        r = 4 + i
        display_cat = cat_label if cat_label else ""
        cell_cat = ws.cell(row=r, column=1, value=display_cat)
        if display_cat:
            cell_cat.font = bold_font
            cell_cat.fill = section_fill
        cell_cat.border = thin_border

        ws.cell(row=r, column=2, value=var_label).font = label_font
        ws.cell(row=r, column=2).border = thin_border

        for j, sname in enumerate(["Conservative", "Base", "Aggressive"]):
            col = 3 + j
            cell = ws.cell(row=r, column=col)
            cell.value = SCENARIOS[sname][key]
            cell.font = label_font
            cell.border = thin_border
            cell.fill = editable_fill
            cell.number_format = fmt

    auto_width(ws, min_w=14)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: 3-YEAR PROJECTION  (Base case, simple formulas with helper rows)
# ═══════════════════════════════════════════════════════════════════════════════
#
# Layout — every formula is at most one operation:
#   Row 3:  Headers
#   Row 4:  INPUTS (section)
#   Row 5:  Units          = Assumptions!$D$4 / $D$5 / $D$6
#   Row 6:  Trucks         = Assumptions!$D$7 / $D$8 / $D$9
#   Row 7:  Rate           = Assumptions!$D$10
#   Row 8:  Event %        = Assumptions!$D$11
#   Row 9:  Event Premium  = Assumptions!$D$12
#   Row 10: (blank)
#   Row 11: REVENUE (section)
#   Row 12: Construction Units  = INT(B5*(1-B8))       ← local refs only
#   Row 13: Event Units         = B5-B12               ← local refs only
#   Row 14: Construction Rev    = B12*B7*12             ← local refs only
#   Row 15: Event Rev           = B13*B9*12             ← local refs only
#   Row 16: Total Revenue       = B14+B15               ← local refs only
#   Row 17: (blank)
#   Row 18: COSTS (section)
#   Row 19: Chemicals      = Assumptions!$D$17 * B5 * 12
#   Row 20: Fuel & Maint   = Assumptions!$D$18 * B6 * 12
#   Row 21: Labor           = Assumptions!$D$13 + Assumptions!$D$14
#   Row 22: Yard            = Assumptions!$D$19 * 12
#   Row 23: Insurance       = Assumptions!$D$20
#   Row 24: Office          = Assumptions!$D$21
#   Row 25: Marketing       = Assumptions!$D$22
#   Row 26: Total Costs     = SUM(B19:B25)
#   Row 27: (blank)
#   Row 28: EBITDA          = B16 - B26                 ← local refs only
#   Row 29: Loan Amount     = Assumptions!$D$24 * (1-Assumptions!$D$27)
#   Row 30: Monthly Payment = -PMT(rate/12, term*12, B29)
#   Row 31: Annual Debt     = B30 * 12                  ← local ref only
#   Row 32: DSCR            = B28 / B31                 ← local ref only
#   Row 33: Net Cash Flow   = B28 - B31                 ← local ref only

def build_projection(ws):
    ws.sheet_properties.tabColor = "27AE60"

    ws.merge_cells("A1:D1")
    ws["A1"].value = "3-Year Financial Projection — Base Case"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"].value = "Gray italic rows are helper inputs. Every formula references local cells where possible."
    ws["A2"].font = note_font

    for i, h in enumerate(["Line Item", "Year 1", "Year 2", "Year 3"], 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 4)

    sc = "D"  # Base case column

    # ── INPUTS section (pull from Assumptions into local cells) ──────────────
    section_row(ws, 4, "INPUTS (from Assumptions)", 4)

    # Row 5: Units
    row_label(ws, 5, "Units", helper=True)
    for yi, col in enumerate([2, 3, 4], 1):
        set_formula(ws, 5, col, f"={a(YEAR_ROWS[yi]['units'], sc)}", fmt=USD, is_helper=True)

    # Row 6: Trucks
    row_label(ws, 6, "Trucks", helper=True)
    for yi, col in enumerate([2, 3, 4], 1):
        set_formula(ws, 6, col, f"={a(YEAR_ROWS[yi]['trucks'], sc)}", fmt=USD, is_helper=True)

    # Row 7: Monthly Rate (same all years)
    row_label(ws, 7, "Monthly Rate", helper=True)
    for col in [2, 3, 4]:
        set_formula(ws, 7, col, f"={a(A_RATE, sc)}", fmt=USD, is_helper=True)

    # Row 8: Event %
    row_label(ws, 8, "Event Mix %", helper=True)
    for col in [2, 3, 4]:
        set_formula(ws, 8, col, f"={a(A_EVENT_PCT, sc)}", fmt=PCT, is_helper=True)

    # Row 9: Event Premium
    row_label(ws, 9, "Event Premium", helper=True)
    for col in [2, 3, 4]:
        set_formula(ws, 9, col, f"={a(A_EVENT_PREMIUM, sc)}", fmt=USD, is_helper=True)

    # Row 10: blank

    # ── REVENUE section (all local references) ──────────────────────────────
    section_row(ws, 11, "REVENUE", 4)

    # Row 12: Construction Units = INT(units * (1 - event%))
    row_label(ws, 12, "  Construction Units", helper=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 12, cn, f"=INT({cl}5*(1-{cl}8))", fmt=USD, is_helper=True)

    # Row 13: Event Units = units - construction units
    row_label(ws, 13, "  Event Units", helper=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 13, cn, f"={cl}5-{cl}12", fmt=USD, is_helper=True)

    # Row 14: Construction Revenue = construction_units * rate * 12
    row_label(ws, 14, "  Construction Revenue")
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 14, cn, f"={cl}12*{cl}7*12", fmt=USD)

    # Row 15: Event Revenue = event_units * event_premium * 12
    row_label(ws, 15, "  Event Revenue")
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 15, cn, f"={cl}13*{cl}9*12", fmt=USD)

    # Row 16: Total Revenue
    row_label(ws, 16, "Total Revenue", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 16, cn, f"={cl}14+{cl}15", fmt=USD, bold=True)

    # Row 17: blank

    # ── COSTS section ────────────────────────────────────────────────────────
    section_row(ws, 18, "COSTS", 4)

    # Row 19: Chemicals = cost_per_unit * units * 12
    row_label(ws, 19, "  Chemicals")
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 19, cn, f"={a(A_CHEM, sc)}*{cl}5*12", fmt=USD)

    # Row 20: Fuel & Maint = cost_per_truck * trucks * 12
    row_label(ws, 20, "  Fuel & Maintenance")
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 20, cn, f"={a(A_FUEL, sc)}*{cl}6*12", fmt=USD)

    # Row 21: Labor = owner + driver (driver is year-specific)
    row_label(ws, 21, "  Labor (Owner + Drivers)")
    for yi, col in enumerate([2, 3, 4], 1):
        set_formula(ws, 21, col,
                    f"={a(A_OWNER, sc)}+{a(YEAR_ROWS[yi]['driver'], sc)}", fmt=USD)

    # Row 22: Yard = monthly * 12
    row_label(ws, 22, "  Yard Lease")
    for col in [2, 3, 4]:
        set_formula(ws, 22, col, f"={a(A_YARD, sc)}*12", fmt=USD)

    # Row 23-25: Annual costs (simple lookups)
    for r, lbl, arow in [(23, "  Insurance", A_INS), (24, "  Office & Misc", A_OFFICE),
                          (25, "  Marketing", A_MKTG)]:
        row_label(ws, r, lbl)
        for col in [2, 3, 4]:
            set_formula(ws, r, col, f"={a(arow, sc)}", fmt=USD)

    # Row 26: Total Costs
    row_label(ws, 26, "Total Costs", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 26, cn, f"=SUM({cl}19:{cl}25)", fmt=USD, bold=True)

    # Row 27: blank

    # ── SUMMARY section (all local references) ──────────────────────────────
    # Row 28: EBITDA
    section_row(ws, 28, "EBITDA", 4)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 28, cn, f"={cl}16-{cl}26", fmt=USD, bold=True)
        ws.cell(row=28, column=cn).fill = section_fill

    # Row 29: Loan Amount (one cross-sheet ref, same all years)
    row_label(ws, 29, "  Loan Amount", helper=True)
    set_formula(ws, 29, 2, f"={a(A_LOAN, sc)}*(1-{a(A_DOWN, sc)})", fmt=USD, is_helper=True)
    for col in [3, 4]:
        set_formula(ws, 29, col, "=B29", fmt=USD, is_helper=True)

    # Row 30: Monthly Payment = -PMT(rate/12, term*12, loan)
    row_label(ws, 30, "  Monthly Payment", helper=True)
    set_formula(ws, 30, 2,
                f"=-PMT({a(A_RATE_SBA, sc)}/12,{a(A_TERM, sc)}*12,B29)", fmt=USD, is_helper=True)
    for col in [3, 4]:
        set_formula(ws, 30, col, "=B30", fmt=USD, is_helper=True)

    # Row 31: Annual Debt Service = monthly * 12
    row_label(ws, 31, "Annual Debt Service")
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 31, cn, f"={cl}30*12", fmt=USD)

    # Row 32: DSCR = EBITDA / Debt
    row_label(ws, 32, "DSCR", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 32, cn, f"={cl}28/{cl}31", fmt=RATIO, bold=True)

    # Row 33: Net Cash Flow = EBITDA - Debt
    section_row(ws, 33, "Net Cash Flow", 4)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 33, cn, f"={cl}28-{cl}31", fmt=USD, bold=True)
        ws.cell(row=33, column=cn).fill = section_fill

    # Conditional formatting
    ws.conditional_formatting.add("B32:D32",
        CellIsRule(operator="greaterThanOrEqual", formula=["1.25"],
                   font=green_font_cf, fill=green_fill_cf))
    ws.conditional_formatting.add("B32:D32",
        CellIsRule(operator="lessThan", formula=["1.25"],
                   font=red_font_cf, fill=red_fill_cf))

    auto_width(ws, min_w=14)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: DASHBOARD  (3-scenario Year 1 — broken into simple individual rows)
# ═══════════════════════════════════════════════════════════════════════════════
#
# Layout — each row is one simple formula:
#   Row 4:  Headers
#   Row 5:  INPUTS
#   Row 6:  Units Y1         = Assumptions!$C/$D/$E$4
#   Row 7:  Trucks Y1        = Assumptions!$C/$D/$E$7
#   Row 8:  Rate             = Assumptions!$C/$D/$E$10
#   Row 9:  Event %          = Assumptions!$C/$D/$E$11
#   Row 10: Event Premium    = Assumptions!$C/$D/$E$12
#   Row 11: (blank)
#   Row 12: REVENUE
#   Row 13: Constr Units     = INT(B6*(1-B9))
#   Row 14: Event Units      = B6-B13
#   Row 15: Constr Rev       = B13*B8*12
#   Row 16: Event Rev        = B14*B10*12
#   Row 17: Total Revenue    = B15+B16
#   Row 18: (blank)
#   Row 19: COSTS
#   Row 20-26: individual cost lines (each one formula)
#   Row 27: Total Costs      = SUM(B20:B26)
#   Row 28: (blank)
#   Row 29: EBITDA           = B17-B27
#   Row 30: Debt Service     = -PMT(...)*12
#   Row 31: DSCR             = B29/B30
#   Row 32: Net Cash         = B29-B30

def build_dashboard(ws):
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:D1")
    ws["A1"].value = "SF Bay Area Portable Sanitation — Financial Model"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"].value = "3-Scenario Comparison (Year 1)"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # Header
    for i, h in enumerate(["Metric", "Conservative", "Base", "Aggressive"], 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 4)

    sc_cols = ["C", "D", "E"]  # Assumptions columns

    # ── INPUTS ───────────────────────────────────────────────────────────────
    section_row(ws, 5, "INPUTS", 4)

    row_label(ws, 6, "Units (Year 1)", helper=True)
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 6, j, f"={a(A_UNITS_Y1, sc)}", fmt=USD, is_helper=True)

    row_label(ws, 7, "Trucks (Year 1)", helper=True)
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 7, j, f"={a(A_TRUCKS_Y1, sc)}", fmt=USD, is_helper=True)

    row_label(ws, 8, "Monthly Rate", helper=True)
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 8, j, f"={a(A_RATE, sc)}", fmt=USD, is_helper=True)

    row_label(ws, 9, "Event Mix %", helper=True)
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 9, j, f"={a(A_EVENT_PCT, sc)}", fmt=PCT, is_helper=True)

    row_label(ws, 10, "Event Premium", helper=True)
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 10, j, f"={a(A_EVENT_PREMIUM, sc)}", fmt=USD, is_helper=True)

    # Row 11: blank

    # ── REVENUE ──────────────────────────────────────────────────────────────
    section_row(ws, 12, "REVENUE", 4)

    row_label(ws, 13, "  Construction Units", helper=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 13, j, f"=INT({cl}6*(1-{cl}9))", fmt=USD, is_helper=True)

    row_label(ws, 14, "  Event Units", helper=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 14, j, f"={cl}6-{cl}13", fmt=USD, is_helper=True)

    row_label(ws, 15, "  Construction Revenue")
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 15, j, f"={cl}13*{cl}8*12", fmt=USD)

    row_label(ws, 16, "  Event Revenue")
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 16, j, f"={cl}14*{cl}10*12", fmt=USD)

    row_label(ws, 17, "Total Revenue", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 17, j, f"={cl}15+{cl}16", fmt=USD, bold=True)

    # Row 18: blank

    # ── COSTS ────────────────────────────────────────────────────────────────
    section_row(ws, 19, "COSTS", 4)

    row_label(ws, 20, "  Chemicals")
    for j, sc in enumerate(sc_cols, 2):
        cl = get_column_letter(j)
        set_formula(ws, 20, j, f"={a(A_CHEM, sc)}*{cl}6*12", fmt=USD)

    row_label(ws, 21, "  Fuel & Maintenance")
    for j, sc in enumerate(sc_cols, 2):
        cl = get_column_letter(j)
        set_formula(ws, 21, j, f"={a(A_FUEL, sc)}*{cl}7*12", fmt=USD)

    row_label(ws, 22, "  Labor")
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 22, j, f"={a(A_OWNER, sc)}+{a(A_DRIVER_Y1, sc)}", fmt=USD)

    row_label(ws, 23, "  Yard Lease")
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 23, j, f"={a(A_YARD, sc)}*12", fmt=USD)

    row_label(ws, 24, "  Insurance")
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 24, j, f"={a(A_INS, sc)}", fmt=USD)

    row_label(ws, 25, "  Office & Misc")
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 25, j, f"={a(A_OFFICE, sc)}", fmt=USD)

    row_label(ws, 26, "  Marketing")
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 26, j, f"={a(A_MKTG, sc)}", fmt=USD)

    row_label(ws, 27, "Total Costs", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 27, j, f"=SUM({cl}20:{cl}26)", fmt=USD, bold=True)

    # Row 28: blank

    # ── SUMMARY ──────────────────────────────────────────────────────────────
    section_row(ws, 29, "EBITDA", 4)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 29, j, f"={cl}17-{cl}27", fmt=USD, bold=True)
        ws.cell(row=29, column=j).fill = section_fill

    row_label(ws, 30, "Annual Debt Service")
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 30, j,
                    f"=-PMT({a(A_RATE_SBA, sc)}/12,{a(A_TERM, sc)}*12,"
                    f"{a(A_LOAN, sc)}*(1-{a(A_DOWN, sc)}))*12", fmt=USD)

    row_label(ws, 31, "DSCR", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 31, j, f"={cl}29/{cl}30", fmt=RATIO, bold=True)

    section_row(ws, 32, "Net Cash Flow", 4)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 32, j, f"={cl}29-{cl}30", fmt=USD, bold=True)
        ws.cell(row=32, column=j).fill = section_fill

    # Takeaway note
    ws.merge_cells("A34:D34")
    ws.cell(row=34, column=1,
            value="All scenarios clear the SBA 1.25x DSCR minimum. "
                  "Edit the Assumptions sheet to update.").font = Font(
        name="Calibri", italic=True, size=11, color=NAVY)

    # Conditional formatting
    ws.conditional_formatting.add("B31:D31",
        CellIsRule(operator="greaterThanOrEqual", formula=["1.25"],
                   font=green_font_cf, fill=green_fill_cf))
    ws.conditional_formatting.add("B31:D31",
        CellIsRule(operator="lessThan", formula=["1.25"],
                   font=red_font_cf, fill=red_fill_cf))
    ws.conditional_formatting.add("B32:D32",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=red_font_cf, fill=red_fill_cf))

    auto_width(ws, min_w=16)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: RETURN ANALYSIS  (references Projection; FV() for loan balance)
# ═══════════════════════════════════════════════════════════════════════════════
def build_returns(ws):
    ws.sheet_properties.tabColor = "8E44AD"

    ws.merge_cells("A1:C1")
    ws["A1"].value = "5-Year Return Analysis — Base Case"
    ws["A1"].font = title_font
    ws.merge_cells("A2:C2")
    ws["A2"].value = "Years 4-5 reuse Year 3 assumptions. All cells are formulas."
    ws["A2"].font = note_font

    for i, h in enumerate(["Metric", "Value", "Notes"], 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 3)

    sc = "D"

    def label(r, text, bold_flag=False):
        row_label(ws, r, text, bold=bold_flag)

    def note(r, text):
        c = ws.cell(row=r, column=3, value=text)
        c.font = note_font
        c.border = thin_border

    # Row 4: Equity = purchase * down%
    label(4, "Equity Invested (Down Payment)")
    set_formula(ws, 4, 2, f"={a(A_PRICE, sc)}*{a(A_DOWN, sc)}", fmt=USD)
    note(4, "Down payment on acquisition")

    # Rows 6-10: Net cash by year (reference Projection row 33)
    for yr, rw in [(1, 6), (2, 7), (3, 8), (4, 9), (5, 10)]:
        label(rw, f"Year {yr} Net Cash Flow")
        if yr <= 3:
            cl = get_column_letter(1 + yr)  # B=Y1, C=Y2, D=Y3
            set_formula(ws, rw, 2, f"={p(f'{cl}33')}", fmt=USD)
        else:
            set_formula(ws, rw, 2, f"={p('D33')}", fmt=USD)
            note(rw, "Uses Year 3 assumptions")

    # Row 12: Cumulative
    label(12, "5-Year Cumulative Cash", bold_flag=True)
    set_formula(ws, 12, 2, "=SUM(B6:B10)", fmt=USD, bold=True)
    note(12, "Total cash distributions to owner")

    # Row 14: Y5 EBITDA
    label(14, "Year 5 EBITDA (SDE Proxy)")
    set_formula(ws, 14, 2, f"={p('D28')}", fmt=USD)
    note(14, "Year 3 EBITDA as proxy")

    # Row 15: Terminal Value = 3x
    label(15, "Terminal Value (3x SDE)", bold_flag=True)
    set_formula(ws, 15, 2, "=B14*3", fmt=USD, bold=True)
    note(15, "Industry standard 2-4x for route businesses")

    # Row 16: Remaining Loan Balance — using FV() instead of manual math
    # FV(rate, nper, pmt, pv) where pv = -loan (we owe it), pmt = payment (outflow)
    label(16, "Remaining Loan Balance (Year 5)")
    set_formula(ws, 16, 2,
                f"=FV({a(A_RATE_SBA, sc)}/12,60,-{p('B30')},-{p('B29')})",
                fmt=USD)
    note(16, "=FV(rate, 60 payments, -pmt, -loan)")

    # Row 17: Equity at Exit
    label(17, "Equity at Exit", bold_flag=True)
    set_formula(ws, 17, 2, "=B15-B16", fmt=USD, bold=True)
    note(17, "Terminal value minus remaining debt")

    # Row 19: Total Return
    label(19, "Total 5-Year Return", bold_flag=True)
    set_formula(ws, 19, 2, "=B12+B17", fmt=USD, bold=True)
    note(19, "Cumulative cash + equity at exit")

    # Row 20: Multiple
    label(20, "Return Multiple", bold_flag=True)
    set_formula(ws, 20, 2, "=B19/B4", fmt=RATIO, bold=True)
    note(20, "Total return on equity invested")

    # Rows 22-28: IRR cash flow series
    row_label(ws, 22, "IRR Cash Flow Series", bold=True)

    label(23, "Year 0 (Investment)")
    set_formula(ws, 23, 2, "=-B4", fmt=USD)

    for yr, src in [(1, 6), (2, 7), (3, 8), (4, 9)]:
        label(23 + yr, f"Year {yr}")
        set_formula(ws, 23 + yr, 2, f"=B{src}", fmt=USD)

    label(28, "Year 5 (incl. exit)")
    set_formula(ws, 28, 2, "=B10+B17", fmt=USD)
    note(28, "Operating cash + equity at exit")

    # Row 30: IRR
    label(30, "Estimated IRR", bold_flag=True)
    set_formula(ws, 30, 2, "=IRR(B23:B28)", fmt=PCT, bold=True)
    note(30, "Excel IRR on cash flow series above")

    # Conditional formatting
    ws.conditional_formatting.add("B30",
        CellIsRule(operator="greaterThan", formula=["0.3"],
                   font=green_font_cf, fill=green_fill_cf))
    ws.conditional_formatting.add("B20",
        CellIsRule(operator="greaterThan", formula=["5"],
                   font=green_font_cf, fill=green_fill_cf))

    auto_width(ws, min_w=16, max_w=50)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()

    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_assumptions = wb.create_sheet("Assumptions")
    ws_proj = wb.create_sheet("3-Year Projection")
    ws_returns = wb.create_sheet("Return Analysis")

    build_assumptions(ws_assumptions)
    build_projection(ws_proj)
    build_dashboard(ws_dash)
    build_returns(ws_returns)

    wb.save(OUT_FILE)
    size = os.path.getsize(OUT_FILE)
    print(f"OK: {OUT_FILE} ({size:,} bytes)")


if __name__ == "__main__":
    main()
