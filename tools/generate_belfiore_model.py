#!/usr/bin/env python3
"""Generate Belfiore Cheese financial model (.xlsx) with live Excel formulas.

Two models in one workbook:
  (A) Business P&L — Belfiore revenue, COGS, operating costs, EBITDA
  (B) Personal Decision — Nick's opportunity cost: traditional career vs. Belfiore paths

Architecture (same pattern as generate_model.py):
  - Assumptions sheet: raw input values (the ONLY place numbers are typed)
  - Business P&L: 3-year projection with formulas referencing Assumptions
  - Dashboard: 3-scenario Year 1 comparison
  - Personal Decision: career path NPV comparison with equity scenarios
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO, "output", "belfiore-cheese")
OUT_FILE = os.path.join(OUT_DIR, "030326-financial-model.xlsx")
ASSUMPTIONS_FILE = os.path.join(OUT_DIR, "assumptions.json")

# ── Styles ───────────────────────────────────────────────────────────────────
NAVY = "1A2744"
GOLD = "C9A84C"
CHEESE_GOLD = "F5DEB3"
GREEN_HEX = "27AE60"
RED_HEX = "E74C3C"

header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
label_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", bold=True, size=11)
title_font = Font(name="Calibri", bold=True, size=16, color=NAVY)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=NAVY)
note_font = Font(name="Calibri", size=10, italic=True, color="666666")
helper_font = Font(name="Calibri", size=10, italic=True, color="888888")
editable_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
section_fill = PatternFill(start_color="E8EAF0", end_color="E8EAF0", fill_type="solid")
helper_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
positive_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
negative_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

USD = '#,##0'
PCT = '0.0%'
RATIO = '0.00"x"'

green_font_cf = Font(name="Calibri", bold=True, color=GREEN_HEX)
red_font_cf = Font(name="Calibri", bold=True, color=RED_HEX)
green_fill_cf = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
red_fill_cf = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")


# ── Assumption row map ───────────────────────────────────────────────────────
# Data starts row 4 on the Assumptions sheet
# Revenue
A_REV_Y1 = 4
A_REV_GROW_Y2 = 5
A_REV_GROW_Y3 = 6
# Revenue mix
A_WHOLESALE_PCT = 7
A_PREMIUM_PCT = 8
A_COPACKING_PCT = 9
A_DTC_PCT = 10
# COGS (as % of revenue)
A_MILK_CONV = 11
A_MILK_ORG = 12
A_MILK_BUF = 13
A_LABOR_PROD = 14
A_PACKAGING = 15
A_UTILITIES = 16
A_DISTRIBUTION = 17
# OpEx (annual fixed costs)
A_OWNER_COMP = 18
A_FACILITY = 19
A_INSURANCE = 20
A_COMPLIANCE = 21
A_MARKETING = 22
A_OTHER_OPEX = 23
# Nick's comp
A_COO_BASE = 24
A_COO_BONUS = 25
A_COO_BENEFITS = 26
# Equity / valuation
A_EQUITY_VEST = 27
A_SDE_MULT = 28
# Alternative career
A_TRAD_COMP = 29
A_TRAD_GROWTH = 30


ASSUMPTION_KEYS = [
    "revenue_y1", "revenue_growth_y2", "revenue_growth_y3",
    "wholesale_pct", "premium_retail_pct", "copacking_pct", "dtc_pct",
    "milk_conventional_pct_rev", "milk_organic_pct_rev", "milk_buffalo_pct_rev",
    "labor_production_pct_rev", "packaging_pct_rev", "utilities_ebmud_pct_rev",
    "distribution_pct_rev",
    "owner_comp", "facility_annual", "insurance_annual", "compliance_annual",
    "marketing_annual", "other_opex_annual",
    "coo_base_salary", "coo_bonus_pct", "coo_benefits_annual",
    "equity_vest_pct_per_year", "sde_multiple",
    "traditional_career_comp_y1", "traditional_career_growth",
]

ASSUMPTION_LABELS = [
    ("Revenue", "Year 1 Revenue ($)"),
    ("", "Year 2 Growth (%)"),
    ("", "Year 3 Growth (%)"),
    ("Revenue Mix", "Wholesale/Retail (%)"),
    ("", "Premium/Specialty Retail (%)"),
    ("", "Co-Packing (%)"),
    ("", "DTC/Online (%)"),
    ("COGS (% of Rev)", "Milk - Conventional (%)"),
    ("", "Milk - Organic Premium (%)"),
    ("", "Milk - Water Buffalo (%)"),
    ("", "Production Labor (%)"),
    ("", "Packaging (%)"),
    ("", "Utilities & EBMUD (%)"),
    ("", "Distribution/Logistics (%)"),
    ("Fixed Costs", "Owner Compensation ($/yr)"),
    ("", "Facility Lease ($/yr)"),
    ("", "Insurance ($/yr)"),
    ("", "Compliance & Permits ($/yr)"),
    ("", "Marketing ($/yr)"),
    ("", "Other OpEx ($/yr)"),
    ("Nick's Comp", "COO Base Salary ($/yr)"),
    ("", "Bonus Target (%)"),
    ("", "Benefits ($/yr)"),
    ("Equity", "Annual Equity Vesting (%)"),
    ("", "SDE Valuation Multiple"),
    ("Alt Career", "Traditional Career Y1 Comp ($)"),
    ("", "Traditional Career Annual Growth (%)"),
]

ASSUMPTION_FMTS = [
    USD, PCT, PCT,
    PCT, PCT, PCT, PCT,
    PCT, PCT, PCT, PCT, PCT, PCT, PCT,
    USD, USD, USD, USD, USD, USD,
    USD, PCT, USD,
    PCT, RATIO,
    USD, PCT,
]


def a(row, col="D"):
    """Absolute reference to Assumptions sheet cell."""
    return f"Assumptions!${col}${row}"


def p(cell):
    """Reference to Business P&L sheet cell."""
    return f"'Business P&&L'!{cell}"


# ── Helpers ──────────────────────────────────────────────────────────────────

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def set_formula(ws, row, col, formula, fmt=None, bold=False, is_helper=False):
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.font = helper_font if is_helper else (bold_font if bold else label_font)
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt
    if is_helper:
        cell.fill = helper_fill
    return cell


def row_label(ws, row, text, bold=False, section=False, helper=False):
    cell = ws.cell(row=row, column=1, value=text)
    if section:
        cell.font = bold_font
        cell.fill = section_fill
    elif helper:
        cell.font = helper_font
        cell.fill = helper_fill
    elif bold:
        cell.font = bold_font
    else:
        cell.font = label_font
    cell.border = thin_border


def section_row(ws, row, text, max_col):
    row_label(ws, row, text, section=True)
    for c in range(2, max_col + 1):
        ws.cell(row=row, column=c).fill = section_fill
        ws.cell(row=row, column=c).border = thin_border


def auto_width(ws, min_w=12, max_w=35):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        best = min_w
        for cell in col_cells:
            if cell.value and not str(cell.value).startswith("="):
                best = max(best, min(len(str(cell.value)) + 3, max_w))
        ws.column_dimensions[letter].width = best


def load_assumptions():
    """Load assumptions.json and return scenario dicts."""
    with open(ASSUMPTIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET: ASSUMPTIONS
# ═════════════════════════════════════════════════════════════════════════════
def build_assumptions(ws, data):
    ws.sheet_properties.tabColor = GOLD

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Belfiore Cheese — Model Assumptions"
    ws["A1"].font = title_font

    ws.merge_cells("A2:E2")
    ws["A2"].value = (
        "Yellow cells are inputs. All other sheets auto-recalculate. "
        "Sources: Agent research (comp, economics, input costs), Gemini analysis."
    )
    ws["A2"].font = note_font

    row = 3
    for i, h in enumerate(["Category", "Variable", "Conservative", "Base", "Aggressive"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 5)

    scenarios = data["scenarios"]
    for i, (key, (cat_label, var_label), fmt) in enumerate(
            zip(ASSUMPTION_KEYS, ASSUMPTION_LABELS, ASSUMPTION_FMTS)):
        r = 4 + i
        display_cat = cat_label if cat_label else ""
        cell_cat = ws.cell(row=r, column=1, value=display_cat)
        if display_cat:
            cell_cat.font = bold_font
            cell_cat.fill = section_fill
        cell_cat.border = thin_border

        ws.cell(row=r, column=2, value=var_label).font = label_font
        ws.cell(row=r, column=2).border = thin_border

        for j, sname in enumerate(["conservative", "base", "aggressive"]):
            col = 3 + j
            cell = ws.cell(row=r, column=col)
            cell.value = scenarios[sname][key]
            cell.font = label_font
            cell.border = thin_border
            cell.fill = editable_fill
            cell.number_format = fmt

    # Input costs reference section
    r = 4 + len(ASSUMPTION_KEYS) + 1
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1).value = ""
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="Reference: Input Cost Data (from Agent 9 research)").font = subtitle_font
    r += 1

    costs = data["input_costs"]
    ref_items = [
        ("Conventional milk", f"${costs['milk_conventional_cwt']:.2f}/cwt", "CDFA Class III CA"),
        ("Organic milk (CCOF)", f"${costs['milk_organic_cwt']:.2f}/cwt", "CCOF certified suppliers"),
        ("Buffalo cream premium", f"{costs['milk_buffalo_premium_multiplier']:.1f}x conventional", "Wilton CA farm"),
        ("Production worker (Berkeley)", f"${costs['berkeley_production_worker_hourly']:.2f}/hr", "BLS + Berkeley min wage"),
        ("Berkeley min wage 2026", f"${costs['berkeley_min_wage_2026']:.2f}/hr", "City of Berkeley"),
        ("CDL driver", f"${costs['cdl_driver_annual']:,}/yr", "BLS Bay Area"),
        ("Lead cheesemaker", f"${costs['cheesemaker_lead_annual']:,}/yr", "Industry benchmarks"),
        ("EBMUD wastewater", f"${costs['ebmud_wastewater_per_ccf']:.2f}/CCF", "EBMUD fee schedule"),
        ("Natural wood smoking", f"${costs['natural_wood_smoking_per_lb']:.2f}/lb cheese", "Industry estimate"),
        ("Cryovac packaging", f"${costs['cryovac_per_unit']:.2f}/unit", "Supplier quotes"),
        ("Water-pack cups", f"${costs['water_pack_cup_per_unit']:.2f}/unit", "Supplier quotes"),
    ]

    for i, (item, value, source) in enumerate(ref_items):
        rr = r + i
        ws.cell(row=rr, column=1, value=item).font = label_font
        ws.cell(row=rr, column=1).border = thin_border
        ws.cell(row=rr, column=2, value=value).font = label_font
        ws.cell(row=rr, column=2).border = thin_border
        ws.cell(row=rr, column=3, value=source).font = note_font
        ws.cell(row=rr, column=3).border = thin_border

    auto_width(ws, min_w=14)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET: BUSINESS P&L (3-Year Projection — Base Case)
# ═════════════════════════════════════════════════════════════════════════════
#
# Layout:
#   Row 3:  Headers
#   Row 4:  REVENUE section
#   Row 5:  Year Revenue helper (Y1 = Assumptions, Y2 = Y1*(1+g), Y3 = Y2*(1+g))
#   Row 6-9: Revenue by channel (helper)
#   Row 10: Total Revenue (= Row 5, bold)
#   Row 11: blank
#   Row 12: COGS section
#   Row 13-19: Individual COGS lines (each = Rev * %)
#   Row 20: Total COGS
#   Row 21: Gross Profit
#   Row 22: Gross Margin %
#   Row 23: blank
#   Row 24: OPERATING EXPENSES section
#   Row 25-30: Fixed costs
#   Row 31: Total OpEx
#   Row 32: blank
#   Row 33: EBITDA section
#   Row 34: EBITDA Margin %
#   Row 35: blank
#   Row 36: OWNER ECONOMICS section
#   Row 37: Owner comp (from assumptions)
#   Row 38: SDE = EBITDA + Owner comp
#   Row 39: blank
#   Row 40: WITH COO section
#   Row 41: COO total comp
#   Row 42: Adjusted EBITDA (EBITDA - COO comp)
#   Row 43: Adjusted Owner comp (reduced)
#   Row 44: New SDE

def build_business_pl(ws):
    ws.sheet_properties.tabColor = GREEN_HEX

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Belfiore Cheese — Business P&L (Base Case)"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"].value = "All formulas reference Assumptions sheet. Gray rows are helper calculations."
    ws["A2"].font = note_font

    for i, h in enumerate(["Line Item", "Year 1", "Year 2", "Year 3"], 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 4)

    sc = "D"  # Base case column on Assumptions

    # ── REVENUE ──────────────────────────────────────────────────────────────
    section_row(ws, 4, "REVENUE", 4)

    # Row 5: Year Revenue
    row_label(ws, 5, "Total Revenue", bold=True)
    # Y1 = direct from assumptions
    set_formula(ws, 5, 2, f"={a(A_REV_Y1, sc)}", fmt=USD, bold=True)
    # Y2 = Y1 * (1 + growth)
    set_formula(ws, 5, 3, f"=B5*(1+{a(A_REV_GROW_Y2, sc)})", fmt=USD, bold=True)
    # Y3 = Y2 * (1 + growth)
    set_formula(ws, 5, 4, f"=C5*(1+{a(A_REV_GROW_Y3, sc)})", fmt=USD, bold=True)

    # Row 6-9: Revenue by channel (helper rows for visibility)
    channels = [
        (6, "  Wholesale/Retail", A_WHOLESALE_PCT),
        (7, "  Premium/Specialty", A_PREMIUM_PCT),
        (8, "  Co-Packing", A_COPACKING_PCT),
        (9, "  DTC/Online", A_DTC_PCT),
    ]
    for r, lbl, a_row in channels:
        row_label(ws, r, lbl, helper=True)
        for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
            set_formula(ws, r, cn, f"={cl}5*{a(a_row, sc)}", fmt=USD, is_helper=True)

    # Row 11: blank

    # ── COGS ─────────────────────────────────────────────────────────────────
    section_row(ws, 11, "COST OF GOODS SOLD", 4)

    cogs_items = [
        (12, "  Milk - Conventional", A_MILK_CONV),
        (13, "  Milk - Organic Premium", A_MILK_ORG),
        (14, "  Milk - Water Buffalo", A_MILK_BUF),
        (15, "  Production Labor", A_LABOR_PROD),
        (16, "  Packaging", A_PACKAGING),
        (17, "  Utilities & EBMUD", A_UTILITIES),
        (18, "  Distribution/Logistics", A_DISTRIBUTION),
    ]
    for r, lbl, a_row in cogs_items:
        row_label(ws, r, lbl)
        for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
            set_formula(ws, r, cn, f"={cl}5*{a(a_row, sc)}", fmt=USD)

    # Row 19: Total COGS
    row_label(ws, 19, "Total COGS", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 19, cn, f"=SUM({cl}12:{cl}18)", fmt=USD, bold=True)

    # Row 20: Gross Profit
    row_label(ws, 20, "Gross Profit", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 20, cn, f"={cl}5-{cl}19", fmt=USD, bold=True)

    # Row 21: Gross Margin %
    row_label(ws, 21, "Gross Margin %", helper=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 21, cn, f"={cl}20/{cl}5", fmt=PCT, is_helper=True)

    # Row 22: blank

    # ── OPERATING EXPENSES ───────────────────────────────────────────────────
    section_row(ws, 23, "OPERATING EXPENSES", 4)

    opex_items = [
        (24, "  Owner Compensation", A_OWNER_COMP),
        (25, "  Facility Lease", A_FACILITY),
        (26, "  Insurance", A_INSURANCE),
        (27, "  Compliance & Permits", A_COMPLIANCE),
        (28, "  Marketing", A_MARKETING),
        (29, "  Other OpEx", A_OTHER_OPEX),
    ]
    for r, lbl, a_row in opex_items:
        row_label(ws, r, lbl)
        for cn in [2, 3, 4]:
            set_formula(ws, r, cn, f"={a(a_row, sc)}", fmt=USD)

    # Row 30: Total OpEx
    row_label(ws, 30, "Total Operating Expenses", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 30, cn, f"=SUM({cl}24:{cl}29)", fmt=USD, bold=True)

    # Row 31: blank

    # ── EBITDA ───────────────────────────────────────────────────────────────
    section_row(ws, 32, "EBITDA (before Nick)", 4)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 32, cn, f"={cl}20-{cl}30", fmt=USD, bold=True)
        ws.cell(row=32, column=cn).fill = section_fill

    row_label(ws, 33, "EBITDA Margin %", helper=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 33, cn, f"={cl}32/{cl}5", fmt=PCT, is_helper=True)

    # Row 34: blank

    # ── SDE (Seller's Discretionary Earnings) ────────────────────────────────
    section_row(ws, 35, "OWNER ECONOMICS", 4)

    row_label(ws, 36, "Owner Compensation")
    for cn in [2, 3, 4]:
        set_formula(ws, 36, cn, f"={a(A_OWNER_COMP, sc)}", fmt=USD)

    row_label(ws, 37, "SDE (EBITDA + Owner Comp)", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 37, cn, f"={cl}32+{cl}36", fmt=USD, bold=True)

    row_label(ws, 38, "Business Valuation (SDE x Multiple)", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 38, cn, f"={cl}37*{a(A_SDE_MULT, sc)}", fmt=USD, bold=True)

    # Row 39: blank

    # ── WITH COO (Nick joins) ────────────────────────────────────────────────
    section_row(ws, 40, "WITH COO (Nick Joins)", 4)

    row_label(ws, 41, "COO Total Compensation")
    for cn in [2, 3, 4]:
        set_formula(ws, 41, cn,
                    f"={a(A_COO_BASE, sc)}*(1+{a(A_COO_BONUS, sc)})+{a(A_COO_BENEFITS, sc)}",
                    fmt=USD)

    row_label(ws, 42, "Adjusted EBITDA (after COO)", bold=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 42, cn, f"={cl}32-{cl}41", fmt=USD, bold=True)

    row_label(ws, 43, "Adjusted EBITDA Margin %", helper=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 43, cn, f"={cl}42/{cl}5", fmt=PCT, is_helper=True)

    row_label(ws, 44, "Owner Take-Home (EBITDA - COO + Owner Comp)")
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 44, cn, f"={cl}42+{cl}36", fmt=USD)

    row_label(ws, 45, "Owner Comp Reduction vs. Status Quo", helper=True)
    for cl, cn in [("B", 2), ("C", 3), ("D", 4)]:
        set_formula(ws, 45, cn, f"={cl}37-{cl}44", fmt=USD, is_helper=True)

    # Conditional formatting on EBITDA margin
    ws.conditional_formatting.add("B33:D33",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.12"],
                   font=green_font_cf, fill=green_fill_cf))
    ws.conditional_formatting.add("B33:D33",
        CellIsRule(operator="lessThan", formula=["0.10"],
                   font=red_font_cf, fill=red_fill_cf))

    auto_width(ws, min_w=14)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET: DASHBOARD (3-Scenario Year 1 Comparison)
# ═════════════════════════════════════════════════════════════════════════════
def build_dashboard(ws):
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Belfiore Cheese — 3-Scenario Comparison (Year 1)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"].value = "Each scenario pulls directly from Assumptions sheet"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    for i, h in enumerate(["Metric", "Conservative", "Base", "Aggressive"], 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 4)

    sc_cols = ["C", "D", "E"]  # Assumptions columns

    # ── Revenue ──────────────────────────────────────────────────────────────
    section_row(ws, 5, "REVENUE", 4)

    row_label(ws, 6, "Year 1 Revenue", bold=True)
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 6, j, f"={a(A_REV_Y1, sc)}", fmt=USD, bold=True)

    # ── COGS ─────────────────────────────────────────────────────────────────
    section_row(ws, 8, "COGS (% of Revenue)", 4)

    cogs_dash = [
        (9, "  Milk (all types)", [A_MILK_CONV, A_MILK_ORG, A_MILK_BUF]),
        (10, "  Production Labor", [A_LABOR_PROD]),
        (11, "  Packaging", [A_PACKAGING]),
        (12, "  Utilities & EBMUD", [A_UTILITIES]),
        (13, "  Distribution", [A_DISTRIBUTION]),
    ]
    for r, lbl, a_rows in cogs_dash:
        row_label(ws, r, lbl)
        for j, sc in enumerate(sc_cols, 2):
            if len(a_rows) > 1:
                parts = "+".join(f"{a(ar, sc)}" for ar in a_rows)
                set_formula(ws, r, j, f"={get_column_letter(j)}6*({parts})", fmt=USD)
            else:
                set_formula(ws, r, j, f"={get_column_letter(j)}6*{a(a_rows[0], sc)}", fmt=USD)

    row_label(ws, 14, "Total COGS", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 14, j, f"=SUM({cl}9:{cl}13)", fmt=USD, bold=True)

    row_label(ws, 15, "Gross Profit", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 15, j, f"={cl}6-{cl}14", fmt=USD, bold=True)

    # ── OpEx ─────────────────────────────────────────────────────────────────
    section_row(ws, 17, "OPERATING EXPENSES", 4)

    opex_dash = [
        (18, "  Owner Comp", A_OWNER_COMP),
        (19, "  Facility", A_FACILITY),
        (20, "  Insurance", A_INSURANCE),
        (21, "  Compliance", A_COMPLIANCE),
        (22, "  Marketing", A_MARKETING),
        (23, "  Other", A_OTHER_OPEX),
    ]
    for r, lbl, ar in opex_dash:
        row_label(ws, r, lbl)
        for j, sc in enumerate(sc_cols, 2):
            set_formula(ws, r, j, f"={a(ar, sc)}", fmt=USD)

    row_label(ws, 24, "Total OpEx", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 24, j, f"=SUM({cl}18:{cl}23)", fmt=USD, bold=True)

    # ── Summary ──────────────────────────────────────────────────────────────
    section_row(ws, 26, "EBITDA (Status Quo)", 4)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 26, j, f"={cl}15-{cl}24", fmt=USD, bold=True)
        ws.cell(row=26, column=j).fill = section_fill

    row_label(ws, 27, "EBITDA Margin %")
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 27, j, f"={cl}26/{cl}6", fmt=PCT)

    row_label(ws, 28, "SDE (EBITDA + Owner Comp)", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 28, j, f"={cl}26+{cl}18", fmt=USD, bold=True)

    # ── With Nick ────────────────────────────────────────────────────────────
    section_row(ws, 30, "WITH COO (Nick Joins)", 4)

    row_label(ws, 31, "COO Total Comp")
    for j, sc in enumerate(sc_cols, 2):
        set_formula(ws, 31, j,
                    f"={a(A_COO_BASE, sc)}*(1+{a(A_COO_BONUS, sc)})+{a(A_COO_BENEFITS, sc)}",
                    fmt=USD)

    row_label(ws, 32, "Adjusted EBITDA", bold=True)
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 32, j, f"={cl}26-{cl}31", fmt=USD, bold=True)

    row_label(ws, 33, "Owner Impact (annual reduction)")
    for j in [2, 3, 4]:
        cl = get_column_letter(j)
        set_formula(ws, 33, j, f"={cl}31", fmt=USD)

    section_row(ws, 35, "BUSINESS VALUATION", 4)

    row_label(ws, 36, "Valuation (SDE x Multiple)", bold=True)
    for j, sc in enumerate(sc_cols, 2):
        cl = get_column_letter(j)
        set_formula(ws, 36, j, f"={cl}28*{a(A_SDE_MULT, sc)}", fmt=USD, bold=True)
        ws.cell(row=36, column=j).fill = section_fill

    # Conditional formatting
    ws.conditional_formatting.add("B27:D27",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.12"],
                   font=green_font_cf, fill=green_fill_cf))
    ws.conditional_formatting.add("B27:D27",
        CellIsRule(operator="lessThan", formula=["0.10"],
                   font=red_font_cf, fill=red_fill_cf))

    auto_width(ws, min_w=16)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET: PERSONAL DECISION MODEL
# ═════════════════════════════════════════════════════════════════════════════
#
# Compares three paths over 5 years:
#   Path A: Traditional career (market-rate comp, annual raises)
#   Path B: Belfiore Bridge (COO comp, no equity, exit year 5)
#   Path C: Belfiore with Equity (COO comp, vesting equity, exit year 5)

def build_personal_decision(ws):
    ws.sheet_properties.tabColor = "8E44AD"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Nick's Personal Decision Model — 5-Year Path Comparison"
    ws["A1"].font = title_font
    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        "Base case assumptions. Change Assumptions sheet to update. "
        "All values in nominal dollars."
    )
    ws["A2"].font = note_font

    for i, h in enumerate(["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"], 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 6)

    sc = "D"  # Base case

    # ── PATH A: Traditional Career ──────────────────────────────────────────
    section_row(ws, 4, "PATH A: Traditional Career", 6)

    row_label(ws, 5, "Annual Compensation", bold=True)
    # Y1
    set_formula(ws, 5, 2, f"={a(A_TRAD_COMP, sc)}", fmt=USD, bold=True)
    # Y2-Y5: prior year * (1 + growth)
    for cn in [3, 4, 5, 6]:
        prev = get_column_letter(cn - 1)
        set_formula(ws, 5, cn,
                    f"={prev}5*(1+{a(A_TRAD_GROWTH, sc)})", fmt=USD, bold=True)

    row_label(ws, 6, "Cumulative Earnings")
    set_formula(ws, 6, 2, "=B5", fmt=USD)
    for cn in [3, 4, 5, 6]:
        prev = get_column_letter(cn - 1)
        cl = get_column_letter(cn)
        set_formula(ws, 6, cn, f"={prev}6+{cl}5", fmt=USD)

    # Row 7: blank

    # ── PATH B: Belfiore Bridge (No Equity) ─────────────────────────────────
    section_row(ws, 8, "PATH B: Belfiore Bridge (No Equity)", 6)

    row_label(ws, 9, "COO Total Compensation", bold=True)
    for cn in [2, 3, 4, 5, 6]:
        set_formula(ws, 9, cn,
                    f"={a(A_COO_BASE, sc)}*(1+{a(A_COO_BONUS, sc)})+{a(A_COO_BENEFITS, sc)}",
                    fmt=USD, bold=True)

    row_label(ws, 10, "Cumulative Earnings")
    set_formula(ws, 10, 2, "=B9", fmt=USD)
    for cn in [3, 4, 5, 6]:
        prev = get_column_letter(cn - 1)
        cl = get_column_letter(cn)
        set_formula(ws, 10, cn, f"={prev}10+{cl}9", fmt=USD)

    row_label(ws, 11, "Equity Value at Exit")
    for cn in [2, 3, 4, 5, 6]:
        set_formula(ws, 11, cn, "=0", fmt=USD)

    row_label(ws, 12, "Total Value (Earnings + Equity)", bold=True)
    for cn in [2, 3, 4, 5, 6]:
        cl = get_column_letter(cn)
        set_formula(ws, 12, cn, f"={cl}10+{cl}11", fmt=USD, bold=True)

    row_label(ws, 13, "Opportunity Cost vs. Path A", helper=True)
    for cn in [2, 3, 4, 5, 6]:
        cl = get_column_letter(cn)
        set_formula(ws, 13, cn, f"={cl}12-{cl}6", fmt=USD, is_helper=True)

    # Row 14: blank

    # ── PATH C: Belfiore with Equity ────────────────────────────────────────
    section_row(ws, 15, "PATH C: Belfiore with Equity Vesting", 6)

    row_label(ws, 16, "COO Total Compensation", bold=True)
    for cn in [2, 3, 4, 5, 6]:
        set_formula(ws, 16, cn,
                    f"={a(A_COO_BASE, sc)}*(1+{a(A_COO_BONUS, sc)})+{a(A_COO_BENEFITS, sc)}",
                    fmt=USD, bold=True)

    row_label(ws, 17, "Cumulative Earnings")
    set_formula(ws, 17, 2, "=B16", fmt=USD)
    for cn in [3, 4, 5, 6]:
        prev = get_column_letter(cn - 1)
        cl = get_column_letter(cn)
        set_formula(ws, 17, cn, f"={prev}17+{cl}16", fmt=USD)

    row_label(ws, 18, "Cumulative Equity % Vested", helper=True)
    for cn in [2, 3, 4, 5, 6]:
        yr = cn - 1  # year number
        set_formula(ws, 18, cn, f"={a(A_EQUITY_VEST, sc)}*{yr}", fmt=PCT, is_helper=True)

    # Equity value = vested % * business valuation
    # Business valuation = SDE * multiple (use Business P&L Y1 SDE as proxy)
    row_label(ws, 19, "Equity Value at Exit", bold=True)
    for cn in [2, 3, 4, 5, 6]:
        cl = get_column_letter(cn)
        # Reference Business P&L row 37 (SDE) col B (Year 1 as proxy) * multiple * vested %
        set_formula(ws, 19, cn,
                    f"={cl}18*'Business P&&L'!B37*{a(A_SDE_MULT, sc)}",
                    fmt=USD, bold=True)

    row_label(ws, 20, "Total Value (Earnings + Equity)", bold=True)
    for cn in [2, 3, 4, 5, 6]:
        cl = get_column_letter(cn)
        set_formula(ws, 20, cn, f"={cl}17+{cl}19", fmt=USD, bold=True)

    row_label(ws, 21, "Opportunity Cost vs. Path A", helper=True)
    for cn in [2, 3, 4, 5, 6]:
        cl = get_column_letter(cn)
        set_formula(ws, 21, cn, f"={cl}20-{cl}6", fmt=USD, is_helper=True)

    # Row 22: blank

    # ── SUMMARY TABLE ────────────────────────────────────────────────────────
    section_row(ws, 23, "5-YEAR SUMMARY", 6)

    ws.cell(row=24, column=1, value="").border = thin_border
    for i, h in enumerate(["Path A: Career", "Path B: Bridge", "Path C: Equity"], 2):
        ws.cell(row=24, column=i, value=h).font = bold_font
        ws.cell(row=24, column=i).border = thin_border

    summary_items = [
        (25, "5-Year Cumulative Earnings", "F6", "F10", "F17"),
        (26, "Equity Value at Year 5", "0", "F11", "F19"),
        (27, "Total 5-Year Value", "F6", "F12", "F20"),
    ]
    for r, lbl, ref_a, ref_b, ref_c in summary_items:
        row_label(ws, r, lbl, bold=(r == 27))
        for j, ref in enumerate([ref_a, ref_b, ref_c], 2):
            if ref == "0":
                set_formula(ws, r, j, "=0", fmt=USD, bold=(r == 27))
            else:
                set_formula(ws, r, j, f"={ref}", fmt=USD, bold=(r == 27))

    row_label(ws, 28, "Opportunity Cost (vs. Path A)", bold=True)
    set_formula(ws, 28, 2, "=0", fmt=USD, bold=True)
    set_formula(ws, 28, 3, "=C27-B27", fmt=USD, bold=True)
    set_formula(ws, 28, 4, "=D27-B27", fmt=USD, bold=True)

    # Row 30: Breakeven note
    ws.merge_cells("A30:F30")
    ws.cell(row=30, column=1,
            value="Negative opportunity cost = Path A earns more. "
                  "Positive = Belfiore path earns more (including equity).").font = note_font

    # ── KEY INSIGHT ──────────────────────────────────────────────────────────
    section_row(ws, 32, "KEY QUESTION", 6)
    ws.merge_cells("A33:F33")
    ws.cell(row=33, column=1,
            value="At what equity vesting rate does Path C break even with Path A? "
                  "Change 'Annual Equity Vesting (%)' on Assumptions sheet until "
                  "opportunity cost = $0.").font = Font(
        name="Calibri", italic=True, size=11, color=NAVY)

    # Conditional formatting on opportunity cost rows
    for rng in ["B13:F13", "B21:F21", "C28:D28"]:
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                       font=green_font_cf, fill=green_fill_cf))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="lessThan", formula=["0"],
                       font=red_font_cf, fill=red_fill_cf))

    auto_width(ws, min_w=14)


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    data = load_assumptions()

    wb = Workbook()

    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_assumptions = wb.create_sheet("Assumptions")
    ws_pl = wb.create_sheet("Business P&L")
    ws_decision = wb.create_sheet("Personal Decision")

    build_assumptions(ws_assumptions, data)
    build_business_pl(ws_pl)
    build_dashboard(ws_dash)
    build_personal_decision(ws_decision)

    wb.save(OUT_FILE)
    size = os.path.getsize(OUT_FILE)
    print(f"OK: {OUT_FILE} ({size:,} bytes)")
    print(f"Sheets: Dashboard, Assumptions, Business P&L, Personal Decision")
    print(f"Source: {ASSUMPTIONS_FILE}")


if __name__ == "__main__":
    main()
